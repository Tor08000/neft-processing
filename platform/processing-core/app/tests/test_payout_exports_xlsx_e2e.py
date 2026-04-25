from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.billing_period import BillingPeriod, BillingPeriodStatus, BillingPeriodType
from app.models.operation import Operation, OperationStatus, OperationType
from app.models.payout_batch import PayoutBatch
from app.services.abac.engine import AbacDecision, AbacEngine
from app.services.decision import DecisionEngine, DecisionOutcome
from app.services.payout_export_xlsx import load_bank_format

from ._money_router_harness import PAYOUT_TEST_TABLES, money_session_context, payout_client_context


class _InMemoryS3Storage:
    _objects: dict[tuple[str, str], bytes] = {}

    def __init__(self, *, bucket: str | None = None):
        self.bucket = bucket or "neft-payouts"

    def put_bytes(self, key: str, payload: bytes, *, content_type: str = "application/octet-stream") -> str:
        self._objects[(self.bucket, key)] = payload
        return f"s3://{self.bucket}/{key}"

    def exists(self, key: str) -> bool:
        return (self.bucket, key) in self._objects

    def get_bytes(self, key: str) -> bytes | None:
        return self._objects.get((self.bucket, key))


@pytest.fixture(autouse=True)
def _stub_policy_and_storage(monkeypatch: pytest.MonkeyPatch):
    from app.api.v1.endpoints import payouts as payouts_api
    from app.services import payout_exports as payout_exports_service

    monkeypatch.setattr(
        DecisionEngine,
        "evaluate",
        lambda *_args, **_kwargs: SimpleNamespace(outcome=DecisionOutcome.ALLOW),
    )
    monkeypatch.setattr(
        AbacEngine,
        "evaluate",
        lambda *_args, **_kwargs: AbacDecision(True, None, [], {"result": True}),
    )
    _InMemoryS3Storage._objects = {}
    monkeypatch.setattr(payout_exports_service, "S3Storage", _InMemoryS3Storage)
    monkeypatch.setattr(payouts_api, "S3Storage", _InMemoryS3Storage)


@pytest.fixture
def session() -> Session:
    with money_session_context(tables=PAYOUT_TEST_TABLES) as db:
        yield db


@pytest.fixture
def admin_client(session: Session) -> TestClient:
    with payout_client_context(db_session=session) as api_client:
        yield api_client


def _seed_captured_operations(db: Session, target_date: date, partner_id: str, count: int = 3) -> None:
    base_dt = datetime.combine(target_date, datetime.min.time()).replace(tzinfo=timezone.utc)
    for idx in range(count):
        amount = 1000 + idx * 100
        db.add(
            Operation(
                ext_operation_id=f"seed-op-{partner_id}-{idx}",
                operation_type=OperationType.COMMIT,
                status=OperationStatus.CAPTURED,
                created_at=base_dt + timedelta(hours=idx + 1),
                updated_at=base_dt + timedelta(hours=idx + 1),
                merchant_id=partner_id,
                terminal_id="terminal-1",
                client_id="client-1",
                card_id="card-1",
                product_id="FUEL",
                amount=amount,
                amount_settled=amount,
                currency="RUB",
                quantity=Decimal("1.0"),
                captured_amount=amount,
                refunded_amount=0,
                response_code="00",
                response_message="OK",
                authorized=True,
            )
        )
    db.commit()


def _seed_billing_period(db: Session, target_date: date, status: BillingPeriodStatus) -> None:
    existing = (
        db.query(BillingPeriod)
        .filter(BillingPeriod.period_type == BillingPeriodType.ADHOC)
        .filter(BillingPeriod.start_at == datetime.combine(target_date, datetime.min.time(), tzinfo=timezone.utc))
        .filter(BillingPeriod.end_at == datetime.combine(target_date, datetime.max.time(), tzinfo=timezone.utc))
        .one_or_none()
    )
    if existing:
        existing.status = status
        db.commit()
        return
    period = BillingPeriod(
        period_type=BillingPeriodType.ADHOC,
        start_at=datetime.combine(target_date, datetime.min.time(), tzinfo=timezone.utc),
        end_at=datetime.combine(target_date, datetime.max.time(), tzinfo=timezone.utc),
        tz="UTC",
        status=status,
    )
    db.add(period)
    db.commit()


def _create_batch(client: TestClient, db: Session, target_date: date, partner_id: str) -> str:
    _seed_captured_operations(db, target_date, partner_id)
    _seed_billing_period(db, target_date, BillingPeriodStatus.FINALIZED)
    payload = {
        "tenant_id": 1,
        "partner_id": partner_id,
        "date_from": target_date.isoformat(),
        "date_to": target_date.isoformat(),
    }
    response = client.post("/api/v1/payouts/close-period", json=payload)
    assert response.status_code == 200
    return response.json()["batch_id"]


def _seed_partner_meta(db: Session, batch_id: str) -> None:
    batch = db.query(PayoutBatch).filter(PayoutBatch.id == batch_id).one()
    existing_meta = batch.meta if isinstance(batch.meta, dict) else {}
    batch.meta = {
        **existing_meta,
        "partner": {
            "name": "OOO Test",
            "inn": "7701234567",
            "bank": {"bik": "044525225", "account": "40702810900000000001"},
        }
    }
    db.commit()


def test_payout_export_xlsx_idempotent(admin_client: TestClient, session: Session):
    target_date = date.today()
    batch_id = _create_batch(admin_client, session, target_date, "partner-xlsx-1")
    _seed_partner_meta(session, batch_id)

    payload = {
        "format": "XLSX",
        "bank_format_code": "SBER_REGISTRY_V1",
        "provider": "bank",
        "external_ref": "BANK-REG-101",
    }
    first = admin_client.post(f"/api/v1/payouts/batches/{batch_id}/export", json=payload)
    assert first.status_code == 200, first.text
    first_body = first.json()

    second = admin_client.post(f"/api/v1/payouts/batches/{batch_id}/export", json=payload)
    assert second.status_code == 200
    second_body = second.json()

    assert first_body["export_id"] == second_body["export_id"]
    assert first_body["object_key"] == second_body["object_key"]
    assert first_body["state"] == "UPLOADED"

    storage = _InMemoryS3Storage(bucket=first_body["bucket"])
    assert storage.exists(first_body["object_key"])


def test_payout_export_xlsx_download_content(admin_client: TestClient, session: Session):
    target_date = date.today()
    batch_id = _create_batch(admin_client, session, target_date, "partner-xlsx-2")
    _seed_partner_meta(session, batch_id)

    payload = {
        "format": "XLSX",
        "bank_format_code": "SBER_REGISTRY_V1",
        "provider": "bank",
        "external_ref": "BANK-REG-102",
    }
    create_resp = admin_client.post(f"/api/v1/payouts/batches/{batch_id}/export", json=payload)
    assert create_resp.status_code == 200, create_resp.text
    export_id = create_resp.json()["export_id"]

    download_resp = admin_client.get(f"/api/v1/payouts/exports/{export_id}/download")
    assert download_resp.status_code == 200
    assert download_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert download_resp.content

    bank_format = load_bank_format("SBER_REGISTRY_V1")
    workbook = load_workbook(filename=BytesIO(download_resp.content))
    worksheet = workbook.active
    assert worksheet.title == bank_format.sheet_name
    headers = [worksheet.cell(row=1, column=idx + 1).value for idx in range(len(bank_format.columns))]
    expected_headers = [column.header for column in bank_format.columns]
    assert headers == expected_headers
    items = session.query(PayoutBatch).filter(PayoutBatch.id == batch_id).one().items
    assert worksheet.max_row == 1 + len(items)


def test_payout_export_xlsx_external_ref_conflict(admin_client: TestClient, session: Session):
    target_date = date.today()
    batch_one = _create_batch(admin_client, session, target_date, "partner-xlsx-3")
    batch_two = _create_batch(admin_client, session, target_date, "partner-xlsx-4")
    _seed_partner_meta(session, batch_two)

    payload_csv = {"format": "CSV", "provider": "bank", "external_ref": "BANK-REG-103"}
    first = admin_client.post(f"/api/v1/payouts/batches/{batch_one}/export", json=payload_csv)
    assert first.status_code == 200

    payload_xlsx = {
        "format": "XLSX",
        "bank_format_code": "SBER_REGISTRY_V1",
        "provider": "bank",
        "external_ref": "BANK-REG-103",
    }
    conflict = admin_client.post(f"/api/v1/payouts/batches/{batch_two}/export", json=payload_xlsx)
    assert conflict.status_code == 409
