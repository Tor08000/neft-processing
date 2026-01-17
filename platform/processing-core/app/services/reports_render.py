from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from typing import Any, Callable, Iterable, Iterator

from sqlalchemy import and_, cast, or_, String
from sqlalchemy.orm import Session

from app.models import (
    Card,
    ClientEmployee,
    ClientUserRole,
    Document,
    DocumentFile,
    DocumentFileType,
    DocumentStatus,
    DocumentType,
    ExportJobReportType,
    FleetDriver,
    FuelCard,
    FuelLimit,
    FuelLimitPeriod,
    FuelLimitScopeType,
    FuelLimitType,
    Operation,
    SupportTicket,
    SupportTicketStatus,
)
from app.models.fleet import EmployeeStatus
from app.models.audit_log import AuditLog

_DOC_TYPE_ALIASES = {
    "CONTRACT": DocumentType.OFFER,
    "OFFER": DocumentType.OFFER,
    "ACT": DocumentType.ACT,
    "INVOICE": DocumentType.INVOICE,
}


class ExportRenderError(Exception):
    pass


class ExportRenderLimitError(ExportRenderError):
    pass


class ExportRenderValidationError(ExportRenderError):
    pass


TOO_MANY_ROWS_ERROR = "too_many_rows_limit_exceeded"


@dataclass(frozen=True)
class ExportRenderResult:
    filename: str
    headers: list[str]
    rows: list[list[object | None]]


@dataclass(frozen=True)
class ExportRenderStreamResult:
    filename: str
    headers: list[str]
    rows: Iterable[list[object | None]]
    estimated_total_rows: int | None = None
    estimated_total_rows: int | None = None


def _parse_date(value: Any, *, field: str) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError as exc:
            raise ExportRenderValidationError(f"invalid_{field}") from exc
    raise ExportRenderValidationError(f"invalid_{field}")


def _parse_int(value: Any, *, field: str) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ExportRenderValidationError(f"invalid_{field}") from exc


def _date_range_bounds(date_from: date | None, date_to: date | None) -> tuple[datetime | None, datetime | None]:
    start = datetime.combine(date_from, time.min) if date_from else None
    end = datetime.combine(date_to, time.max) if date_to else None
    return start, end


def _extract_token_tail(value: str | None) -> str | None:
    if not value:
        return None
    digits = "".join(char for char in value if char.isdigit())
    if digits:
        return digits[-4:]
    return value[-4:] if len(value) >= 4 else value


def _limits_summary(limits: list[FuelLimit]) -> str | None:
    if not limits:
        return None
    parts: list[str] = []
    for limit in limits:
        period = limit.period.value if isinstance(limit.period, FuelLimitPeriod) else str(limit.period)
        limit_type = limit.limit_type.value if isinstance(limit.limit_type, FuelLimitType) else str(limit.limit_type)
        value = limit.value
        suffix = ""
        if limit.limit_type == FuelLimitType.VOLUME:
            suffix = " L"
        elif limit.currency:
            suffix = f" {limit.currency}"
        parts.append(f"{limit_type}/{period}: {value}{suffix}")
    return "; ".join(parts)


def _format_csv_value(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _build_filename(report_type: ExportJobReportType) -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return f"{report_type.value}_export_{stamp}.csv"


def _stream_query(query, *, chunk_size: int) -> Iterator:
    return query.execution_options(stream_results=True).yield_per(chunk_size)


def _chunked(iterator: Iterable[Any], *, chunk_size: int) -> Iterator[list[Any]]:
    chunk: list[Any] = []
    for item in iterator:
        chunk.append(item)
        if len(chunk) >= chunk_size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


def normalize_filters(report_type: ExportJobReportType, filters: dict[str, Any]) -> dict[str, Any]:
    raw_filters = filters or {}
    normalized: dict[str, Any] = {}
    if report_type == ExportJobReportType.CARDS:
        normalized["status"] = raw_filters.get("status") or None
        normalized["driver_id"] = raw_filters.get("driver_id") or None
        from_date = _parse_date(raw_filters.get("from"), field="from")
        to_date = _parse_date(raw_filters.get("to"), field="to")
        normalized["from"] = from_date.isoformat() if from_date else None
        normalized["to"] = to_date.isoformat() if to_date else None
        return normalized
    if report_type == ExportJobReportType.USERS:
        normalized["role"] = raw_filters.get("role") or None
        normalized["status"] = raw_filters.get("status") or None
        from_date = _parse_date(raw_filters.get("from"), field="from")
        to_date = _parse_date(raw_filters.get("to"), field="to")
        normalized["from"] = from_date.isoformat() if from_date else None
        normalized["to"] = to_date.isoformat() if to_date else None
        return normalized
    if report_type == ExportJobReportType.TRANSACTIONS:
        date_from = _parse_date(raw_filters.get("from"), field="from")
        date_to = _parse_date(raw_filters.get("to"), field="to")
        if not date_from or not date_to:
            raise ExportRenderValidationError("date_range_required")
        normalized["from"] = date_from.isoformat()
        normalized["to"] = date_to.isoformat()
        normalized["status"] = raw_filters.get("status") or None
        normalized["card_id"] = raw_filters.get("card_id") or None
        card_ids = raw_filters.get("card_ids") or []
        if isinstance(card_ids, str):
            card_ids = [item.strip() for item in card_ids.split(",") if item.strip()]
        normalized["card_ids"] = card_ids
        normalized["min_amount"] = _parse_int(raw_filters.get("min_amount"), field="min_amount")
        normalized["max_amount"] = _parse_int(raw_filters.get("max_amount"), field="max_amount")
        return normalized
    if report_type == ExportJobReportType.DOCUMENTS:
        normalized["type"] = raw_filters.get("type") or None
        normalized["status"] = raw_filters.get("status") or None
        from_date = _parse_date(raw_filters.get("from"), field="from")
        to_date = _parse_date(raw_filters.get("to"), field="to")
        normalized["from"] = from_date.isoformat() if from_date else None
        normalized["to"] = to_date.isoformat() if to_date else None
        return normalized
    if report_type == ExportJobReportType.AUDIT:
        from_dt = raw_filters.get("from")
        to_dt = raw_filters.get("to")
        if isinstance(from_dt, (date, datetime)):
            from_dt = from_dt.isoformat()
        if isinstance(to_dt, (date, datetime)):
            to_dt = to_dt.isoformat()
        normalized["from"] = from_dt or None
        normalized["to"] = to_dt or None
        normalized["action"] = raw_filters.get("action") or None
        normalized["actor"] = raw_filters.get("actor") or None
        normalized["entity_type"] = raw_filters.get("entity_type") or None
        normalized["entity_id"] = raw_filters.get("entity_id") or None
        normalized["request_id"] = raw_filters.get("request_id") or None
        return normalized
    if report_type == ExportJobReportType.SUPPORT:
        normalized["status"] = raw_filters.get("status") or None
        from_dt = raw_filters.get("from")
        to_dt = raw_filters.get("to")
        if isinstance(from_dt, (date, datetime)):
            from_dt = from_dt.isoformat()
        if isinstance(to_dt, (date, datetime)):
            to_dt = to_dt.isoformat()
        normalized["from"] = from_dt or None
        normalized["to"] = to_dt or None
        return normalized
    raise ExportRenderValidationError("unsupported_report_type")


def render_cards_csv(db: Session, *, client_id: str, filters: dict[str, Any]) -> ExportRenderResult:
    status = filters.get("status")
    driver_id = filters.get("driver_id")
    date_from = _parse_date(filters.get("from"), field="from")
    date_to = _parse_date(filters.get("to"), field="to")
    start, end = _date_range_bounds(date_from, date_to)
    query = db.query(FuelCard).filter(FuelCard.client_id == str(client_id))
    if status:
        query = query.filter(FuelCard.status == status.upper().strip())
    if driver_id:
        query = query.filter(FuelCard.driver_id == driver_id)
    if start:
        query = query.filter(FuelCard.created_at >= start)
    if end:
        query = query.filter(FuelCard.created_at <= end)

    cards = query.order_by(FuelCard.created_at.desc(), FuelCard.id.desc()).all()

    driver_ids = {str(card.driver_id) for card in cards if card.driver_id}
    drivers = db.query(FleetDriver).filter(FleetDriver.id.in_(driver_ids)).all() if driver_ids else []
    driver_map = {str(driver.id): driver for driver in drivers}

    card_ids = [str(card.id) for card in cards]
    limits = []
    if card_ids:
        limits = (
            db.query(FuelLimit)
            .filter(FuelLimit.client_id == str(client_id))
            .filter(FuelLimit.scope_type == FuelLimitScopeType.CARD)
            .filter(FuelLimit.scope_id.in_(card_ids))
            .filter(FuelLimit.active.is_(True))
            .order_by(FuelLimit.created_at.desc())
            .all()
        )
    limit_map: dict[str, list[FuelLimit]] = {}
    for item in limits:
        if item.scope_id:
            limit_map.setdefault(str(item.scope_id), []).append(item)

    rows = []
    for card in cards:
        driver = driver_map.get(str(card.driver_id)) if card.driver_id else None
        assigned_driver = driver.full_name if driver else None
        rows.append(
            [
                str(card.id),
                card.masked_pan,
                _extract_token_tail(card.masked_pan or card.token_ref),
                card.status.value if hasattr(card.status, "value") else str(card.status),
                assigned_driver,
                _limits_summary(limit_map.get(str(card.id), [])),
                card.created_at,
            ]
        )

    return ExportRenderResult(
        filename=_build_filename(ExportJobReportType.CARDS),
        headers=["card_id", "masked_pan", "token_tail", "status", "assigned_driver", "limit_summary", "created_at"],
        rows=rows,
    )


def render_cards_xlsx(db: Session, *, client_id: str, filters: dict[str, Any]) -> bytes:
    result = render_cards_csv(db, client_id=client_id, filters=filters)
    return render_xlsx_payload(result)


def render_users_csv(db: Session, *, client_id: str, filters: dict[str, Any]) -> ExportRenderResult:
    role = filters.get("role")
    status = filters.get("status")
    date_from = _parse_date(filters.get("from"), field="from")
    date_to = _parse_date(filters.get("to"), field="to")
    start, end = _date_range_bounds(date_from, date_to)
    query = (
        db.query(ClientEmployee)
        .outerjoin(
            ClientUserRole,
            and_(
                ClientUserRole.client_id == str(client_id),
                ClientUserRole.user_id == ClientEmployee.id,
            ),
        )
        .filter(ClientEmployee.client_id == str(client_id))
    )
    if status:
        try:
            parsed_status = EmployeeStatus(status.upper().strip())
        except ValueError as exc:
            raise ExportRenderValidationError("invalid_status") from exc
        query = query.filter(ClientEmployee.status == parsed_status)
    if role:
        role_value = str(role).upper().strip()
        if role_value == "CLIENT_USER":
            query = query.filter(or_(ClientUserRole.roles.ilike(f"%{role_value}%"), ClientUserRole.roles.is_(None)))
        else:
            query = query.filter(ClientUserRole.roles.ilike(f"%{role_value}%"))
    if start:
        query = query.filter(ClientEmployee.created_at >= start)
    if end:
        query = query.filter(ClientEmployee.created_at <= end)

    users = query.order_by(ClientEmployee.created_at.desc(), ClientEmployee.id.desc()).all()

    user_ids = [str(user_item.id) for user_item in users]
    role_rows = (
        db.query(ClientUserRole)
        .filter(ClientUserRole.client_id == str(client_id), ClientUserRole.user_id.in_(user_ids))
        .all()
    )
    role_map = {row.user_id: row.roles for row in role_rows}

    rows = [
        [
            str(user_item.id),
            user_item.email,
            role_map.get(str(user_item.id), "CLIENT_USER"),
            user_item.status.value if user_item.status else None,
            user_item.created_at,
            None,
        ]
        for user_item in users
    ]

    return ExportRenderResult(
        filename=_build_filename(ExportJobReportType.USERS),
        headers=["user_id", "email", "roles", "status", "created_at", "last_login_at"],
        rows=rows,
    )


def render_users_xlsx(db: Session, *, client_id: str, filters: dict[str, Any]) -> bytes:
    result = render_users_csv(db, client_id=client_id, filters=filters)
    return render_xlsx_payload(result)


def render_transactions_csv(db: Session, *, client_id: str, filters: dict[str, Any]) -> ExportRenderResult:
    date_from = _parse_date(filters.get("from"), field="from")
    date_to = _parse_date(filters.get("to"), field="to")
    if not date_from or not date_to:
        raise ExportRenderValidationError("date_range_required")
    start, end = _date_range_bounds(date_from, date_to)
    status = filters.get("status")
    card_id = filters.get("card_id")
    card_ids = filters.get("card_ids") or []
    min_amount = _parse_int(filters.get("min_amount"), field="min_amount")
    max_amount = _parse_int(filters.get("max_amount"), field="max_amount")
    query = db.query(Operation).filter(Operation.client_id == str(client_id))
    if card_id:
        query = query.filter(Operation.card_id == card_id)
    if card_ids:
        query = query.filter(Operation.card_id.in_(card_ids))
    if status:
        query = query.filter(Operation.status == status)
    if start:
        query = query.filter(Operation.created_at >= start)
    if end:
        query = query.filter(Operation.created_at <= end)
    if min_amount is not None:
        query = query.filter(Operation.amount >= min_amount)
    if max_amount is not None:
        query = query.filter(Operation.amount <= max_amount)

    operations = query.order_by(Operation.created_at.asc(), Operation.id.asc()).all()

    card_ids_map = {op.card_id for op in operations}
    cards = db.query(Card).filter(Card.id.in_(card_ids_map)).all() if card_ids_map else []
    card_map = {card.id: card for card in cards}

    rows = [
        [
            str(op.operation_id),
            op.card_id,
            card_map.get(op.card_id).pan_masked if op.card_id in card_map else None,
            op.created_at,
            op.amount,
            op.currency,
            op.product_type.value if hasattr(op.product_type, "value") else op.product_type,
            op.merchant_id,
            op.terminal_id,
            op.status.value if hasattr(op.status, "value") else op.status,
        ]
        for op in operations
    ]

    return ExportRenderResult(
        filename=_build_filename(ExportJobReportType.TRANSACTIONS),
        headers=[
            "transaction_id",
            "card_id",
            "masked_pan",
            "date",
            "amount",
            "currency",
            "product_type",
            "station",
            "network",
            "status",
        ],
        rows=rows,
    )


def render_transactions_xlsx(db: Session, *, client_id: str, filters: dict[str, Any]) -> bytes:
    result = render_transactions_csv(db, client_id=client_id, filters=filters)
    return render_xlsx_payload(result)


def render_documents_csv(db: Session, *, client_id: str, filters: dict[str, Any]) -> ExportRenderResult:
    document_type = filters.get("type")
    status = filters.get("status")
    date_from = _parse_date(filters.get("from"), field="from")
    date_to = _parse_date(filters.get("to"), field="to")
    query = db.query(Document).filter(Document.client_id == str(client_id))
    if date_from:
        query = query.filter(Document.period_from >= date_from)
    if date_to:
        query = query.filter(Document.period_to <= date_to)
    if document_type:
        resolved_type = _DOC_TYPE_ALIASES.get(str(document_type).upper().strip())
        if resolved_type is None:
            try:
                resolved_type = DocumentType(document_type)
            except ValueError as exc:
                raise ExportRenderValidationError("invalid_document_type") from exc
        query = query.filter(Document.document_type == resolved_type)
    if status:
        try:
            parsed_status = DocumentStatus(status)
        except ValueError as exc:
            raise ExportRenderValidationError("invalid_document_status") from exc
        query = query.filter(Document.status == parsed_status)

    documents = query.order_by(Document.period_to.desc(), Document.id.desc()).all()

    document_ids = [str(item.id) for item in documents]
    files = (
        db.query(DocumentFile)
        .filter(DocumentFile.document_id.in_(document_ids))
        .filter(DocumentFile.file_type == DocumentFileType.PDF)
        .all()
        if document_ids
        else []
    )
    file_map = {str(item.document_id): item for item in files}

    rows = []
    for item in documents:
        meta = item.meta if isinstance(item.meta, dict) else {}
        amount = None
        currency = None
        for key in ("amount_total", "total_amount", "amount"):
            if key in meta:
                amount = meta.get(key)
                break
        if isinstance(meta, dict):
            currency = meta.get("currency")
        file_item = file_map.get(str(item.id))
        file_name = file_item.object_key.split("/")[-1] if file_item else None
        rows.append(
            [
                str(item.id),
                item.document_type.value,
                item.number,
                item.period_to,
                item.status.value,
                amount,
                currency,
                file_name,
            ]
        )

    return ExportRenderResult(
        filename=_build_filename(ExportJobReportType.DOCUMENTS),
        headers=["document_id", "type", "number", "date", "status", "amount", "currency", "file_name"],
        rows=rows,
    )


def render_audit_csv(
    db: Session,
    *,
    tenant_id: int,
    allowed_entity_types: set[str] | None,
    filters: dict[str, Any],
) -> ExportRenderResult:
    from_dt = filters.get("from")
    to_dt = filters.get("to")
    action = filters.get("action")
    actor = filters.get("actor")
    entity_type = filters.get("entity_type")
    entity_id = filters.get("entity_id")
    request_id = filters.get("request_id")

    query = db.query(AuditLog).filter(AuditLog.tenant_id == tenant_id)
    if allowed_entity_types is not None:
        if entity_type and entity_type not in allowed_entity_types:
            query = query.filter(AuditLog.id.is_(None))
        else:
            query = query.filter(AuditLog.entity_type.in_(allowed_entity_types))
    if from_dt:
        query = query.filter(AuditLog.ts >= datetime.fromisoformat(from_dt))
    if to_dt:
        query = query.filter(AuditLog.ts <= datetime.fromisoformat(to_dt))
    if action:
        action_list = action if isinstance(action, list) else [action]
        query = query.filter(AuditLog.action.in_(action_list))
    if actor:
        like = f"%{actor}%"
        query = query.filter(or_(AuditLog.actor_email.ilike(like), AuditLog.actor_id.ilike(like)))
    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)
    if entity_id:
        entity_like = f"%{entity_id}%"
        external_refs = cast(AuditLog.external_refs, String)
        query = query.filter(
            or_(AuditLog.entity_id == entity_id, AuditLog.entity_id.ilike(entity_like), external_refs.ilike(entity_like))
        )
    if request_id:
        query = query.filter(AuditLog.request_id == request_id)

    logs = query.order_by(AuditLog.ts.desc(), AuditLog.id.desc()).all()

    rows = []
    for log in logs:
        refs = log.external_refs if isinstance(log.external_refs, dict) else {}
        entity_label = None
        for key in ("masked_pan", "card_masked_pan", "card_tail", "pan_tail", "token_tail", "label", "number_tail"):
            value = refs.get(key)
            if value:
                entity_label = str(value)
                break
        rows.append(
            [
                str(log.id),
                log.ts,
                str(tenant_id),
                log.actor_id or "",
                log.actor_email or log.actor_id or "",
                log.action or "",
                log.entity_type or "",
                log.entity_id or "",
                entity_label or "",
                log.request_id or "",
                str(log.ip) if log.ip else "",
                log.user_agent or "",
                "",
                log.reason or log.action or log.event_type or "",
            ]
        )

    return ExportRenderResult(
        filename=_build_filename(ExportJobReportType.AUDIT),
        headers=[
            "id",
            "created_at",
            "org_id",
            "actor_user_id",
            "actor_label",
            "action",
            "entity_type",
            "entity_id",
            "entity_label",
            "request_id",
            "ip",
            "ua",
            "result",
            "summary",
        ],
        rows=rows,
    )


def render_support_csv(
    db: Session,
    *,
    org_id: str,
    created_by_user_id: str | None,
    filters: dict[str, Any],
) -> ExportRenderResult:
    status = filters.get("status")
    from_dt = filters.get("from")
    to_dt = filters.get("to")
    query = db.query(SupportTicket).filter(SupportTicket.org_id == org_id)
    if status:
        try:
            parsed_status = SupportTicketStatus(status)
        except ValueError as exc:
            raise ExportRenderValidationError("invalid_status") from exc
        query = query.filter(SupportTicket.status == parsed_status)
    if created_by_user_id:
        query = query.filter(SupportTicket.created_by_user_id == created_by_user_id)
    if from_dt:
        query = query.filter(SupportTicket.created_at >= datetime.fromisoformat(from_dt))
    if to_dt:
        query = query.filter(SupportTicket.created_at <= datetime.fromisoformat(to_dt))

    tickets = query.order_by(SupportTicket.created_at.desc(), SupportTicket.id.desc()).all()


def stream_cards_csv(
    db: Session,
    *,
    client_id: str,
    filters: dict[str, Any],
    max_rows: int,
    chunk_size: int,
) -> ExportRenderStreamResult:
    status = filters.get("status")
    driver_id = filters.get("driver_id")
    date_from = _parse_date(filters.get("from"), field="from")
    date_to = _parse_date(filters.get("to"), field="to")
    start, end = _date_range_bounds(date_from, date_to)
    query = db.query(FuelCard).filter(FuelCard.client_id == str(client_id))
    if status:
        query = query.filter(FuelCard.status == status.upper().strip())
    if driver_id:
        query = query.filter(FuelCard.driver_id == driver_id)
    if start:
        query = query.filter(FuelCard.created_at >= start)
    if end:
        query = query.filter(FuelCard.created_at <= end)
    query = query.order_by(FuelCard.created_at.desc(), FuelCard.id.desc()).limit(max_rows + 1)

    def row_iterator() -> Iterator[list[object | None]]:
        streamed = _stream_query(query, chunk_size=chunk_size)
        for chunk in _chunked(streamed, chunk_size=chunk_size):
            driver_ids = {str(card.driver_id) for card in chunk if card.driver_id}
            drivers = db.query(FleetDriver).filter(FleetDriver.id.in_(driver_ids)).all() if driver_ids else []
            driver_map = {str(driver.id): driver for driver in drivers}

            card_ids = [str(card.id) for card in chunk]
            limits: list[FuelLimit] = []
            if card_ids:
                limits = (
                    db.query(FuelLimit)
                    .filter(FuelLimit.client_id == str(client_id))
                    .filter(FuelLimit.scope_type == FuelLimitScopeType.CARD)
                    .filter(FuelLimit.scope_id.in_(card_ids))
                    .filter(FuelLimit.active.is_(True))
                    .order_by(FuelLimit.created_at.desc())
                    .all()
                )
            limit_map: dict[str, list[FuelLimit]] = {}
            for item in limits:
                if item.scope_id:
                    limit_map.setdefault(str(item.scope_id), []).append(item)

            for card in chunk:
                driver = driver_map.get(str(card.driver_id)) if card.driver_id else None
                assigned_driver = driver.full_name if driver else None
                yield [
                    str(card.id),
                    card.masked_pan,
                    _extract_token_tail(card.masked_pan or card.token_ref),
                    card.status.value if hasattr(card.status, "value") else str(card.status),
                    assigned_driver,
                    _limits_summary(limit_map.get(str(card.id), [])),
                    card.created_at,
                ]

    return ExportRenderStreamResult(
        filename=_build_filename(ExportJobReportType.CARDS),
        headers=["card_id", "masked_pan", "token_tail", "status", "assigned_driver", "limit_summary", "created_at"],
        rows=row_iterator(),
    )


def stream_users_csv(
    db: Session,
    *,
    client_id: str,
    filters: dict[str, Any],
    max_rows: int,
    chunk_size: int,
) -> ExportRenderStreamResult:
    role = filters.get("role")
    status = filters.get("status")
    date_from = _parse_date(filters.get("from"), field="from")
    date_to = _parse_date(filters.get("to"), field="to")
    start, end = _date_range_bounds(date_from, date_to)
    query = (
        db.query(ClientEmployee)
        .outerjoin(
            ClientUserRole,
            and_(
                ClientUserRole.client_id == str(client_id),
                ClientUserRole.user_id == ClientEmployee.id,
            ),
        )
        .filter(ClientEmployee.client_id == str(client_id))
    )
    if status:
        try:
            parsed_status = EmployeeStatus(status.upper().strip())
        except ValueError as exc:
            raise ExportRenderValidationError("invalid_status") from exc
        query = query.filter(ClientEmployee.status == parsed_status)
    if role:
        role_value = str(role).upper().strip()
        if role_value == "CLIENT_USER":
            query = query.filter(or_(ClientUserRole.roles.ilike(f"%{role_value}%"), ClientUserRole.roles.is_(None)))
        else:
            query = query.filter(ClientUserRole.roles.ilike(f"%{role_value}%"))
    if start:
        query = query.filter(ClientEmployee.created_at >= start)
    if end:
        query = query.filter(ClientEmployee.created_at <= end)
    query = query.order_by(ClientEmployee.created_at.desc(), ClientEmployee.id.desc()).limit(max_rows + 1)

    def row_iterator() -> Iterator[list[object | None]]:
        streamed = _stream_query(query, chunk_size=chunk_size)
        for chunk in _chunked(streamed, chunk_size=chunk_size):
            user_ids = [str(user_item.id) for user_item in chunk]
            role_rows = (
                db.query(ClientUserRole)
                .filter(ClientUserRole.client_id == str(client_id), ClientUserRole.user_id.in_(user_ids))
                .all()
            )
            role_map = {row.user_id: row.roles for row in role_rows}

            for user_item in chunk:
                yield [
                    str(user_item.id),
                    user_item.email,
                    role_map.get(str(user_item.id), "CLIENT_USER"),
                    user_item.status.value if user_item.status else None,
                    user_item.created_at,
                    None,
                ]

    return ExportRenderStreamResult(
        filename=_build_filename(ExportJobReportType.USERS),
        headers=["user_id", "email", "roles", "status", "created_at", "last_login_at"],
        rows=row_iterator(),
    )


def stream_transactions_csv(
    db: Session,
    *,
    client_id: str,
    filters: dict[str, Any],
    max_rows: int,
    chunk_size: int,
) -> ExportRenderStreamResult:
    date_from = _parse_date(filters.get("from"), field="from")
    date_to = _parse_date(filters.get("to"), field="to")
    if not date_from or not date_to:
        raise ExportRenderValidationError("date_range_required")
    start, end = _date_range_bounds(date_from, date_to)
    status = filters.get("status")
    card_id = filters.get("card_id")
    card_ids = filters.get("card_ids") or []
    min_amount = _parse_int(filters.get("min_amount"), field="min_amount")
    max_amount = _parse_int(filters.get("max_amount"), field="max_amount")
    base_query = db.query(Operation).filter(Operation.client_id == str(client_id))
    if card_id:
        base_query = base_query.filter(Operation.card_id == card_id)
    if card_ids:
        base_query = base_query.filter(Operation.card_id.in_(card_ids))
    if status:
        base_query = base_query.filter(Operation.status == status)
    if start:
        base_query = base_query.filter(Operation.created_at >= start)
    if end:
        base_query = base_query.filter(Operation.created_at <= end)
    if min_amount is not None:
        base_query = base_query.filter(Operation.amount >= min_amount)
    if max_amount is not None:
        base_query = base_query.filter(Operation.amount <= max_amount)
    estimated_total_rows = base_query.count()
    query = base_query.order_by(Operation.created_at.asc(), Operation.id.asc()).limit(max_rows + 1)

    def row_iterator() -> Iterator[list[object | None]]:
        streamed = _stream_query(query, chunk_size=chunk_size)
        for chunk in _chunked(streamed, chunk_size=chunk_size):
            card_ids_map = {op.card_id for op in chunk}
            cards = db.query(Card).filter(Card.id.in_(card_ids_map)).all() if card_ids_map else []
            card_map = {card.id: card for card in cards}

            for op in chunk:
                yield [
                    str(op.operation_id),
                    op.card_id,
                    card_map.get(op.card_id).pan_masked if op.card_id in card_map else None,
                    op.created_at,
                    op.amount,
                    op.currency,
                    op.product_type.value if hasattr(op.product_type, "value") else op.product_type,
                    op.merchant_id,
                    op.terminal_id,
                    op.status.value if hasattr(op.status, "value") else op.status,
                ]

    return ExportRenderStreamResult(
        filename=_build_filename(ExportJobReportType.TRANSACTIONS),
        headers=[
            "transaction_id",
            "card_id",
            "masked_pan",
            "date",
            "amount",
            "currency",
            "product_type",
            "station",
            "network",
            "status",
        ],
        rows=row_iterator(),
        estimated_total_rows=estimated_total_rows,
    )


def stream_documents_csv(
    db: Session,
    *,
    client_id: str,
    filters: dict[str, Any],
    max_rows: int,
    chunk_size: int,
) -> ExportRenderStreamResult:
    document_type = filters.get("type")
    status = filters.get("status")
    date_from = _parse_date(filters.get("from"), field="from")
    date_to = _parse_date(filters.get("to"), field="to")
    query = db.query(Document).filter(Document.client_id == str(client_id))
    if date_from:
        query = query.filter(Document.period_from >= date_from)
    if date_to:
        query = query.filter(Document.period_to <= date_to)
    if document_type:
        resolved_type = _DOC_TYPE_ALIASES.get(str(document_type).upper().strip())
        if resolved_type is None:
            try:
                resolved_type = DocumentType(document_type)
            except ValueError as exc:
                raise ExportRenderValidationError("invalid_document_type") from exc
        query = query.filter(Document.document_type == resolved_type)
    if status:
        try:
            parsed_status = DocumentStatus(status)
        except ValueError as exc:
            raise ExportRenderValidationError("invalid_document_status") from exc
        query = query.filter(Document.status == parsed_status)
    query = query.order_by(Document.period_to.desc(), Document.id.desc()).limit(max_rows + 1)

    def row_iterator() -> Iterator[list[object | None]]:
        streamed = _stream_query(query, chunk_size=chunk_size)
        for chunk in _chunked(streamed, chunk_size=chunk_size):
            document_ids = [str(item.id) for item in chunk]
            files = (
                db.query(DocumentFile)
                .filter(DocumentFile.document_id.in_(document_ids))
                .filter(DocumentFile.file_type == DocumentFileType.PDF)
                .all()
                if document_ids
                else []
            )
            file_map = {str(item.document_id): item for item in files}

            for item in chunk:
                meta = item.meta if isinstance(item.meta, dict) else {}
                amount = None
                currency = None
                for key in ("amount_total", "total_amount", "amount"):
                    if key in meta:
                        amount = meta.get(key)
                        break
                if isinstance(meta, dict):
                    currency = meta.get("currency")
                file_item = file_map.get(str(item.id))
                file_name = file_item.object_key.split("/")[-1] if file_item else None
                yield [
                    str(item.id),
                    item.document_type.value,
                    item.number,
                    item.period_to,
                    item.status.value,
                    amount,
                    currency,
                    file_name,
                ]

    return ExportRenderStreamResult(
        filename=_build_filename(ExportJobReportType.DOCUMENTS),
        headers=["document_id", "type", "number", "date", "status", "amount", "currency", "file_name"],
        rows=row_iterator(),
    )


def stream_audit_csv(
    db: Session,
    *,
    tenant_id: int,
    allowed_entity_types: set[str] | None,
    filters: dict[str, Any],
    max_rows: int,
    chunk_size: int,
) -> ExportRenderStreamResult:
    from_dt = filters.get("from")
    to_dt = filters.get("to")
    action = filters.get("action")
    actor = filters.get("actor")
    entity_type = filters.get("entity_type")
    entity_id = filters.get("entity_id")
    request_id = filters.get("request_id")

    query = db.query(AuditLog).filter(AuditLog.tenant_id == tenant_id)
    if allowed_entity_types is not None:
        if entity_type and entity_type not in allowed_entity_types:
            query = query.filter(AuditLog.id.is_(None))
        else:
            query = query.filter(AuditLog.entity_type.in_(allowed_entity_types))
    if from_dt:
        query = query.filter(AuditLog.ts >= datetime.fromisoformat(from_dt))
    if to_dt:
        query = query.filter(AuditLog.ts <= datetime.fromisoformat(to_dt))
    if action:
        action_list = action if isinstance(action, list) else [action]
        query = query.filter(AuditLog.action.in_(action_list))
    if actor:
        like = f"%{actor}%"
        query = query.filter(or_(AuditLog.actor_email.ilike(like), AuditLog.actor_id.ilike(like)))
    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)
    if entity_id:
        entity_like = f"%{entity_id}%"
        external_refs = cast(AuditLog.external_refs, String)
        query = query.filter(
            or_(
                AuditLog.entity_id == entity_id,
                AuditLog.entity_id.ilike(entity_like),
                external_refs.ilike(entity_like),
            )
        )
    if request_id:
        query = query.filter(AuditLog.request_id == request_id)
    query = query.order_by(AuditLog.ts.desc(), AuditLog.id.desc()).limit(max_rows + 1)

    def row_iterator() -> Iterator[list[object | None]]:
        streamed = _stream_query(query, chunk_size=chunk_size)
        for chunk in _chunked(streamed, chunk_size=chunk_size):
            for log in chunk:
                refs = log.external_refs if isinstance(log.external_refs, dict) else {}
                entity_label = None
                for key in ("masked_pan", "card_masked_pan", "card_tail", "pan_tail", "token_tail", "label", "number_tail"):
                    value = refs.get(key)
                    if value:
                        entity_label = str(value)
                        break
                yield [
                    str(log.id),
                    log.ts,
                    str(tenant_id),
                    log.actor_id or "",
                    log.actor_email or log.actor_id or "",
                    log.action or "",
                    log.entity_type or "",
                    log.entity_id or "",
                    entity_label or "",
                    log.request_id or "",
                    str(log.ip) if log.ip else "",
                    log.user_agent or "",
                    "",
                    log.reason or log.action or log.event_type or "",
                ]

    return ExportRenderStreamResult(
        filename=_build_filename(ExportJobReportType.AUDIT),
        headers=[
            "id",
            "created_at",
            "org_id",
            "actor_user_id",
            "actor_label",
            "action",
            "entity_type",
            "entity_id",
            "entity_label",
            "request_id",
            "ip",
            "ua",
            "result",
            "summary",
        ],
        rows=row_iterator(),
    )


def stream_support_csv(
    db: Session,
    *,
    org_id: str,
    created_by_user_id: str | None,
    filters: dict[str, Any],
    max_rows: int,
    chunk_size: int,
) -> ExportRenderStreamResult:
    status = filters.get("status")
    from_dt = filters.get("from")
    to_dt = filters.get("to")
    query = db.query(SupportTicket).filter(SupportTicket.org_id == org_id)
    if status:
        try:
            parsed_status = SupportTicketStatus(status)
        except ValueError as exc:
            raise ExportRenderValidationError("invalid_status") from exc
        query = query.filter(SupportTicket.status == parsed_status)
    if created_by_user_id:
        query = query.filter(SupportTicket.created_by_user_id == created_by_user_id)
    if from_dt:
        query = query.filter(SupportTicket.created_at >= datetime.fromisoformat(from_dt))
    if to_dt:
        query = query.filter(SupportTicket.created_at <= datetime.fromisoformat(to_dt))
    query = query.order_by(SupportTicket.created_at.desc(), SupportTicket.id.desc()).limit(max_rows + 1)

    def row_iterator() -> Iterator[list[object | None]]:
        streamed = _stream_query(query, chunk_size=chunk_size)
        for chunk in _chunked(streamed, chunk_size=chunk_size):
            for ticket in chunk:
                yield [
                    str(ticket.id),
                    ticket.created_at,
                    ticket.updated_at,
                    ticket.status.value,
                    ticket.priority.value,
                    ticket.subject,
                    ticket.created_by_user_id,
                    ticket.sla_first_response_status.value,
                    ticket.sla_resolution_status.value,
                ]

    return ExportRenderStreamResult(
        filename=_build_filename(ExportJobReportType.SUPPORT),
        headers=[
            "ticket_id",
            "created_at",
            "updated_at",
            "status",
            "priority",
            "subject",
            "created_by_user_id",
            "sla_first_response_status",
            "sla_resolution_status",
        ],
        rows=row_iterator(),
    )

    rows = [
        [
            str(ticket.id),
            ticket.created_at,
            ticket.updated_at,
            ticket.status.value,
            ticket.priority.value,
            ticket.subject,
            ticket.created_by_user_id,
            ticket.sla_first_response_status.value,
            ticket.sla_resolution_status.value,
        ]
        for ticket in tickets
    ]

    return ExportRenderResult(
        filename=_build_filename(ExportJobReportType.SUPPORT),
        headers=[
            "ticket_id",
            "created_at",
            "updated_at",
            "status",
            "priority",
            "subject",
            "created_by_user_id",
            "sla_first_response_status",
            "sla_resolution_status",
        ],
        rows=rows,
    )


def render_export_report(
    db: Session,
    *,
    report_type: ExportJobReportType,
    client_id: str | None,
    tenant_id: int | None,
    org_id: str | None,
    created_by_user_id: str | None,
    filters: dict[str, Any],
    allowed_entity_types: set[str] | None = None,
) -> ExportRenderResult:
    if report_type == ExportJobReportType.CARDS:
        if not client_id:
            raise ExportRenderValidationError("missing_client")
        return render_cards_csv(db, client_id=client_id, filters=filters)
    if report_type == ExportJobReportType.USERS:
        if not client_id:
            raise ExportRenderValidationError("missing_client")
        return render_users_csv(db, client_id=client_id, filters=filters)
    if report_type == ExportJobReportType.TRANSACTIONS:
        if not client_id:
            raise ExportRenderValidationError("missing_client")
        return render_transactions_csv(db, client_id=client_id, filters=filters)
    if report_type == ExportJobReportType.DOCUMENTS:
        if not client_id:
            raise ExportRenderValidationError("missing_client")
        return render_documents_csv(db, client_id=client_id, filters=filters)
    if report_type == ExportJobReportType.AUDIT:
        if tenant_id is None:
            raise ExportRenderValidationError("missing_tenant")
        return render_audit_csv(db, tenant_id=tenant_id, allowed_entity_types=allowed_entity_types, filters=filters)
    if report_type == ExportJobReportType.SUPPORT:
        if not org_id:
            raise ExportRenderValidationError("missing_org")
        filter_user_id = filters.get("created_by_user_id") if isinstance(filters, dict) else None
        return render_support_csv(
            db,
            org_id=org_id,
            created_by_user_id=filter_user_id or created_by_user_id,
            filters=filters,
        )
    raise ExportRenderValidationError("unsupported_report_type")


def render_export_report_stream(
    db: Session,
    *,
    report_type: ExportJobReportType,
    client_id: str | None,
    tenant_id: int | None,
    org_id: str | None,
    created_by_user_id: str | None,
    filters: dict[str, Any],
    max_rows: int,
    chunk_size: int,
    allowed_entity_types: set[str] | None = None,
) -> ExportRenderStreamResult:
    if report_type == ExportJobReportType.CARDS:
        if not client_id:
            raise ExportRenderValidationError("missing_client")
        return stream_cards_csv(db, client_id=client_id, filters=filters, max_rows=max_rows, chunk_size=chunk_size)
    if report_type == ExportJobReportType.USERS:
        if not client_id:
            raise ExportRenderValidationError("missing_client")
        return stream_users_csv(db, client_id=client_id, filters=filters, max_rows=max_rows, chunk_size=chunk_size)
    if report_type == ExportJobReportType.TRANSACTIONS:
        if not client_id:
            raise ExportRenderValidationError("missing_client")
        return stream_transactions_csv(
            db, client_id=client_id, filters=filters, max_rows=max_rows, chunk_size=chunk_size
        )
    if report_type == ExportJobReportType.DOCUMENTS:
        if not client_id:
            raise ExportRenderValidationError("missing_client")
        return stream_documents_csv(db, client_id=client_id, filters=filters, max_rows=max_rows, chunk_size=chunk_size)
    if report_type == ExportJobReportType.AUDIT:
        if tenant_id is None:
            raise ExportRenderValidationError("missing_tenant")
        return stream_audit_csv(
            db,
            tenant_id=tenant_id,
            allowed_entity_types=allowed_entity_types,
            filters=filters,
            max_rows=max_rows,
            chunk_size=chunk_size,
        )
    if report_type == ExportJobReportType.SUPPORT:
        if not org_id:
            raise ExportRenderValidationError("missing_org")
        filter_user_id = filters.get("created_by_user_id") if isinstance(filters, dict) else None
        return stream_support_csv(
            db,
            org_id=org_id,
            created_by_user_id=filter_user_id or created_by_user_id,
            filters=filters,
            max_rows=max_rows,
            chunk_size=chunk_size,
        )
    raise ExportRenderValidationError("unsupported_report_type")


def render_csv_payload(result: ExportRenderResult) -> bytes:
    from io import StringIO

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(result.headers)
    for row in result.rows:
        writer.writerow([_format_csv_value(value) for value in row])
    return buffer.getvalue().encode("utf-8")


def _xlsx_cell_value(value: object | None) -> object | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value
    return value


def _xlsx_display_text(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def render_xlsx_payload(result: ExportRenderResult) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(result.headers)

    for row in result.rows:
        sheet.append([_xlsx_cell_value(value) for value in row])

    column_widths = [len(header) if header else 0 for header in result.headers]
    for row in result.rows:
        for idx, value in enumerate(row):
            column_widths[idx] = max(column_widths[idx], len(_xlsx_display_text(value)))

    for idx, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(idx)
        sheet.column_dimensions[column_letter].width = max(10, min(width + 2, 60))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_csv_stream(
    result: ExportRenderStreamResult,
    *,
    file_path: str,
    max_rows: int,
    progress_callback: Callable[[int], None] | None = None,
) -> int:
    row_count = 0
    with open(file_path, "w", newline="", encoding="utf-8") as file_handle:
        writer = csv.writer(file_handle)
        writer.writerow(result.headers)
        for row in result.rows:
            row_count += 1
            if row_count > max_rows:
                raise ExportRenderLimitError(TOO_MANY_ROWS_ERROR)
            writer.writerow([_format_csv_value(value) for value in row])
            if progress_callback:
                progress_callback(row_count)
    return row_count


def write_xlsx_stream(
    result: ExportRenderStreamResult,
    *,
    file_path: str,
    max_rows: int,
    progress_callback: Callable[[int], None] | None = None,
) -> int:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Sheet1")
    sheet.append(result.headers)
    for idx, header in enumerate(result.headers, start=1):
        column_letter = get_column_letter(idx)
        sheet.column_dimensions[column_letter].width = max(10, min(len(str(header)) + 2, 60))

    row_count = 0
    for row in result.rows:
        row_count += 1
        if row_count > max_rows:
            raise ExportRenderLimitError(TOO_MANY_ROWS_ERROR)
        sheet.append([_xlsx_cell_value(value) for value in row])
        if progress_callback:
            progress_callback(row_count)

    workbook.save(file_path)
    return row_count


__all__ = [
    "ExportRenderError",
    "ExportRenderLimitError",
    "ExportRenderResult",
    "ExportRenderStreamResult",
    "ExportRenderValidationError",
    "TOO_MANY_ROWS_ERROR",
    "normalize_filters",
    "render_cards_xlsx",
    "render_csv_payload",
    "render_export_report",
    "render_export_report_stream",
    "render_transactions_xlsx",
    "render_users_xlsx",
    "render_xlsx_payload",
    "write_csv_stream",
    "write_xlsx_stream",
]
