from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from functools import lru_cache
from io import BytesIO
from pathlib import Path
import re
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session, selectinload

from app.models.payout_batch import PayoutBatch


BANK_FORMATS_DIR = Path(__file__).resolve().parent.parent / "bank_formats"
MONEY_FORMAT = "#,##0.00"
DATE_FORMAT = "DD.MM.YYYY"


@dataclass(frozen=True)
class BankFormatColumn:
    key: str
    header: str
    width: int | None
    type: str
    source: str
    transform: tuple[str, ...]


@dataclass(frozen=True)
class BankFormat:
    code: str
    title: str
    sheet_name: str
    filename_pattern: str
    columns: tuple[BankFormatColumn, ...]


@dataclass(frozen=True)
class PayoutXlsxResult:
    payload: bytes
    meta: dict[str, Any]


def _load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        content = yaml.safe_load(handle)
    if not isinstance(content, dict):
        raise ValueError(f"invalid bank format file: {path}")
    return content


def _parse_bank_format(data: dict[str, Any]) -> BankFormat:
    columns_raw = data.get("columns", [])
    if not isinstance(columns_raw, list):
        raise ValueError("columns must be a list")
    columns: list[BankFormatColumn] = []
    for column in columns_raw:
        if not isinstance(column, dict):
            raise ValueError("column must be a mapping")
        transform = column.get("transform") or []
        if transform is None:
            transform = []
        if not isinstance(transform, list):
            raise ValueError("transform must be a list")
        columns.append(
            BankFormatColumn(
                key=str(column["key"]),
                header=str(column["header"]),
                width=int(column.get("width")) if column.get("width") is not None else None,
                type=str(column.get("type") or "text"),
                source=str(column["source"]),
                transform=tuple(str(item) for item in transform),
            )
        )
    return BankFormat(
        code=str(data["code"]),
        title=str(data.get("title", data["code"])),
        sheet_name=str(data.get("sheet_name", "Registry")),
        filename_pattern=str(data.get("filename_pattern", "payout_{partner_id}_{date_from}_{date_to}.xlsx")),
        columns=tuple(columns),
    )


@lru_cache
def _bank_formats_index() -> dict[str, BankFormat]:
    formats: dict[str, BankFormat] = {}
    for path in BANK_FORMATS_DIR.glob("*.yaml"):
        data = _load_yaml(path)
        bank_format = _parse_bank_format(data)
        formats[bank_format.code] = bank_format
    return formats


def list_bank_formats() -> list[dict[str, str]]:
    formats = _bank_formats_index().values()
    return [
        {"code": bank_format.code, "title": bank_format.title}
        for bank_format in sorted(formats, key=lambda fmt: fmt.code)
    ]


def load_bank_format(code: str) -> BankFormat:
    bank_format = _bank_formats_index().get(code)
    if not bank_format:
        raise ValueError("bank_format_not_found")
    return bank_format


def _digits_only(value: Any) -> str | None:
    if value is None:
        return None
    return re.sub(r"\D+", "", str(value))


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _format_date(value: Any, fmt: str) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime(fmt)
    return str(value)


def _apply_transform(value: Any, transform: str) -> Any:
    if transform.startswith("default:"):
        default_value = transform.split(":", 1)[1]
        if value in (None, "", []):
            return default_value
        return value
    if transform == "digits_only":
        return _digits_only(value)
    if transform == "upper":
        return str(value).upper() if value is not None else None
    if transform == "trim":
        return str(value).strip() if value is not None else None
    if transform == "money_2dp":
        dec = _to_decimal(value)
        return dec.quantize(Decimal("0.01")) if dec is not None else None
    if transform == "date_iso":
        if value is None:
            return None
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        return str(value)
    if transform == "date_ddmmyyyy":
        return _format_date(value, "%d.%m.%Y")
    return value


def _resolve_source(
    source: str,
    *,
    batch: PayoutBatch,
    item: Any,
    partner_meta: dict[str, Any],
    computed: dict[str, Any],
    export_context: dict[str, Any],
) -> Any:
    parts = source.split(".")
    if not parts:
        return None
    root = parts[0]
    path = parts[1:]
    if root == "batch":
        value = batch
    elif root == "item":
        value = item
    elif root == "partner":
        value = partner_meta
    elif root == "computed":
        value = computed
    elif root == "export":
        value = export_context
    else:
        return None

    for key in path:
        if value is None:
            return None
        if isinstance(value, dict):
            value = value.get(key)
        else:
            value = getattr(value, key, None)
    return value


def _compute_values(batch: PayoutBatch) -> dict[str, Any]:
    period = f"{batch.date_from.isoformat()} - {batch.date_to.isoformat()}"
    purpose = f"Payouts for {batch.partner_id} {batch.date_from.isoformat()} - {batch.date_to.isoformat()}"
    return {
        "period": period,
        "payment_purpose": purpose,
    }


def generate_payout_registry_xlsx(
    db: Session,
    *,
    batch_id: str,
    format_code: str,
    provider: str | None,
    external_ref: str | None,
) -> PayoutXlsxResult:
    batch = (
        db.query(PayoutBatch)
        .options(selectinload(PayoutBatch.items))
        .filter(PayoutBatch.id == batch_id)
        .one_or_none()
    )
    if not batch:
        raise ValueError("batch_not_found")

    bank_format = load_bank_format(format_code)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = bank_format.sheet_name

    header_font = Font(bold=True)
    for column_index, column in enumerate(bank_format.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column.header)
        cell.font = header_font
        if column.width:
            worksheet.column_dimensions[get_column_letter(column_index)].width = column.width

    partner_meta = {}
    if isinstance(batch.meta, dict):
        partner_meta = batch.meta.get("partner") or {}

    export_context = {"external_ref": external_ref, "provider": provider}
    computed = _compute_values(batch)
    for row_index, item in enumerate(batch.items or [], start=2):
        for column_index, column in enumerate(bank_format.columns, start=1):
            value = _resolve_source(
                column.source,
                batch=batch,
                item=item,
                partner_meta=partner_meta,
                computed=computed,
                export_context=export_context,
            )
            for transform in column.transform:
                value = _apply_transform(value, transform)
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if column.type == "money":
                cell.number_format = MONEY_FORMAT
            elif column.type == "date":
                cell.number_format = DATE_FORMAT

    stream = BytesIO()
    workbook.save(stream)
    payload = stream.getvalue()
    meta = {
        "bank_format_code": bank_format.code,
        "sheet_name": bank_format.sheet_name,
        "columns": [column.key for column in bank_format.columns],
    }
    return PayoutXlsxResult(payload=payload, meta=meta)


def build_filename(
    *,
    format_code: str | None,
    partner_id: str,
    date_from: date,
    date_to: date,
    external_ref: str | None,
) -> str:
    if not format_code:
        return f"payout_{partner_id}_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
    bank_format = load_bank_format(format_code)
    ref_value = external_ref or "no-ref"
    filename = bank_format.filename_pattern
    return (
        filename.replace("{partner_id}", partner_id)
        .replace("{date_from}", date_from.isoformat())
        .replace("{date_to}", date_to.isoformat())
        .replace("{external_ref}", ref_value)
    )


__all__ = [
    "PayoutXlsxResult",
    "build_filename",
    "generate_payout_registry_xlsx",
    "list_bank_formats",
    "load_bank_format",
]
